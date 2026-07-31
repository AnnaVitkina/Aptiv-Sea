"""Load accessorial charge tabs and append them to the output workbook."""

import math
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ACCESSORIAL_OUTPUT_COLS = [
    "Source tab",
    "Rate cost name",
    "Rate Card Name",
    "Multiplier",
    "Rate by",
    "Currency",
    "Price",
    "Max",
    "Apply if",
]

FLOW_ACCESSORIAL_CONFIG = {
    "1": {"multiplier": ""},
    "2": {"multiplier": "Quantity/Percentage"},
    "3": {"multiplier": "Quantity/Percentage"},
}

CHARGE_HEADER_ALIASES = ("charge item", "charge heads")
LANE_HEADER_ALIASES = ("lane id",)
MEASUREMENT_HEADER_ALIASES = ("unit price", "unit / currency", "unit currency")
COMMENT_HEADER_ALIASES = ("comment", "comments", "remarks", "note", "notes")

GENERIC_UNIT_VALUES = {
    "per entry",
    "per fcr",
    "per trk move/hour",
    "per trk move/30 mins",
    "per booking",
    "per so / booking per time",
    "per document",
    "contianer per hr",
    "per shipment",
    "per fcr/hbl/cbl",
}

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FONT_HEADER = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
ALIGN_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
ALIGN_LEFT_WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")

ACCESSORIAL_COLUMN_WIDTHS = {
    "Source tab": 28,
    "Rate cost name": 42,
    "Rate Card Name": 42,
    "Multiplier": 18,
    "Rate by": 22,
    "Currency": 10,
    "Price": 24,
    "Max": 12,
    "Apply if": 36,
}


def normalize(s: str) -> str:
    return re.sub(r"[-_]", " ", str(s).strip().lower())


def normalize_quotes(s: str) -> str:
    return str(s).replace("´", "'").replace("`", "'").replace("'", "'").strip()


def cell_matches_aliases(value, aliases: tuple[str, ...]) -> bool:
    if pd.isna(value):
        return False
    n = normalize(value)
    return any(n == alias or alias in n for alias in aliases)


def find_col(columns: list[str], target: str) -> str | None:
    t = normalize(target)
    for c in columns:
        if normalize(c) == t:
            return c
    return None


def find_col_by_aliases(columns: list[str], aliases: tuple[str, ...]) -> str | None:
    for c in columns:
        n = normalize(c)
        for alias in aliases:
            if n == alias or alias in n:
                return c
    return None


def round_up(value, decimals=3):
    if pd.isna(value):
        return value
    try:
        multiplier = 10 ** decimals
        return math.ceil(float(value) * multiplier) / multiplier
    except (TypeError, ValueError):
        return value


def find_accessorial_header_row(raw: pd.DataFrame, max_rows: int = 50) -> int | None:
    for row_idx in range(min(len(raw), max_rows)):
        for col_idx in range(len(raw.columns)):
            if cell_matches_aliases(raw.iloc[row_idx, col_idx], CHARGE_HEADER_ALIASES):
                return row_idx
    return None


def read_accessorial_sheet(xlsx: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(xlsx, sheet_name=sheet_name, header=None)
    header_row = find_accessorial_header_row(raw)
    if header_row is None:
        print(f"  Warning: header row not found in '{sheet_name}' — skipping tab")
        return pd.DataFrame()

    headers = []
    for col_idx, value in enumerate(raw.iloc[header_row]):
        if pd.isna(value) or str(value).strip() == "":
            headers.append(f"Unnamed: {col_idx}")
        else:
            headers.append(str(value).strip())

    df = raw.iloc[header_row + 1:].copy()
    df.columns = headers
    return df.dropna(how="all").reset_index(drop=True)


def find_amount_currency_columns(columns: list[str]) -> tuple[str | None, str | None]:
    combined = find_col(columns, "Currency and Amount")
    if combined:
        return combined, None

    for col in columns:
        name = str(col).strip().upper()
        if re.fullmatch(r"[A-Z]{3}", name):
            return col, name

    return None, None


def parse_currency_amount(value, default_currency: str | None = None) -> tuple[str, object]:
    if pd.isna(value):
        return default_currency or "", ""

    text = str(value).strip()
    if not text:
        return default_currency or "", ""

    thereafter_match = re.search(
        r"thereafter\s+(?:(USD|EUR|[A-Z]{3})\s*)?[$€]?\s*([\d.,]+)",
        text,
        re.IGNORECASE,
    )
    if thereafter_match:
        currency = (thereafter_match.group(1) or "").upper()
        if not currency and "$" in thereafter_match.group(0):
            currency = "USD"
        return currency or default_currency or "", round_up(thereafter_match.group(2).replace(",", ""))

    match = re.match(r"^([\d.,]+)\s*([A-Za-z]{3})$", text)
    if match:
        return match.group(2).upper(), round_up(match.group(1).replace(",", ""))

    match = re.match(r"^([A-Za-z]{3})\s*([\d.,]+)$", text)
    if match:
        return match.group(1).upper(), round_up(match.group(2).replace(",", ""))

    try:
        return default_currency or "", round_up(float(text.replace(",", "")))
    except ValueError:
        return default_currency or "", text


def extract_max_amount(value, default_currency: str | None = None) -> tuple[str, object]:
    """Extract maximum charge cap from descriptive price text."""
    if pd.isna(value):
        return default_currency or "", ""

    text = str(value).strip()
    if not text:
        return default_currency or "", ""

    max_match = re.search(
        r"maximum charge of\s*(?:(USD|EUR|[A-Z]{3})\s*)?[$€]?\s*([\d.,]+)",
        text,
        re.IGNORECASE,
    )
    if not max_match:
        max_match = re.search(
            r"(?:max(?:imum)?\s+(?:charge|cap)|cap set at).*?[$€]\s*([\d.,]+)",
            text,
            re.IGNORECASE,
        )
        if max_match:
            amount = max_match.group(1)
            currency = "USD" if "$" in max_match.group(0) else (default_currency or "")
            return currency, round_up(amount.replace(",", ""))
    if max_match:
        currency = (max_match.group(1) or "").upper()
        if not currency and "$" in max_match.group(0):
            currency = "USD"
        elif not currency and "€" in max_match.group(0):
            currency = "EUR"
        return currency or default_currency or "", round_up(max_match.group(2).replace(",", ""))

    return default_currency or "", ""


def extract_max_from_entry(entry: dict, default_currency: str | None = None) -> tuple[str, object]:
    texts = [entry.get("price_raw"), *entry.get("row_texts", [])]
    for text in texts:
        currency, max_val = extract_max_amount(text, default_currency)
        if max_val != "":
            return currency, max_val
    return default_currency or "", ""


def row_cell_texts(row: pd.Series, columns: list[str]) -> list[str]:
    texts: list[str] = []
    for col in columns:
        value = row.get(col)
        if pd.isna(value):
            continue
        text = str(value).strip()
        if text and text.lower() != "nan":
            texts.append(text)
    return texts


def row_has_hamida_comment(row: pd.Series, columns: list[str]) -> bool:
    return any(is_hamida_case_by_case_text(row.get(col)) for col in columns)


def split_lane_ids(value) -> list[str]:
    if pd.isna(value):
        return [""]
    parts = re.split(r"[,;/]+", str(value).strip())
    return [part.strip() for part in parts if part.strip()]


def is_skippable_charge(value) -> bool:
    if pd.isna(value):
        return True
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return True
    if text.upper().startswith("IF APPLICABLE"):
        return True
    return is_hamida_case_by_case_text(text)


def is_hamida_case_by_case_text(value) -> bool:
    if pd.isna(value):
        return False
    text = str(value).strip().lower()
    if not text:
        return False
    if "hamida" in text:
        return True
    if "charge to aptiv" in text and "case by case" in text:
        return True
    if "charge to aptiv" in text and "approve" in text:
        return True
    return False


def is_skippable_row(charge, comment=None) -> bool:
    if is_skippable_charge(charge):
        return True
    if comment is not None and is_hamida_case_by_case_text(comment):
        return True
    return False


def extract_tab_country(sheet_name: str) -> str:
    match = re.search(r"Charges\s+(\w+)\s+Lane", sheet_name, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return sheet_name


def split_container_measurements(measurement: str) -> list[str]:
    """Split combined container values such as 40' DRY / 40' HC into separate types."""
    text = normalize_quotes(measurement)
    if not text:
        return [""]

    if re.search(r"40\s*'\s*DRY\s*/\s*40\s*'\s*HC", text, re.IGNORECASE):
        return ["40' DRY", "40' HC"]

    return [text]


def is_generic_per_container(measurement: str) -> bool:
    m = normalize(measurement)
    if m in ("per container", "container"):
        return True
    if "per container" in m:
        return True
    return False


CONTAINER_TYPE_RE = re.compile(
    r"(?:^|\s)(?:20|40)\s*['\u2019]\s*(?:dry|hc|ft)?(?:\s|$)",
    re.IGNORECASE,
)
VALID_DIRECT_TRK_CONTAINER_RE = re.compile(
    r"^(?:20|40)\s*['\u2019]\s*(?:dry|hc|ft)\s*$",
    re.IGNORECASE,
)


def is_pricing_or_tiered_text(text: str) -> bool:
    t = normalize(str(text).strip())
    if not t:
        return False
    if is_tiered_container_price(text):
        return True
    if re.search(r"\bdays?\s+\d+\s+to\s+\d+\b", t):
        return True
    if "$" in t or "€" in t:
        return True
    if re.search(r"\bper\s+(?:20|40)\b", t):
        return True
    if "onwards" in t:
        return True
    if re.search(r"\bor\s+40\b", t):
        return True
    return False


def is_valid_direct_trk_container(text: str) -> bool:
    candidate = normalize_quotes(str(text)).strip()
    if not candidate or is_pricing_or_tiered_text(candidate):
        return False
    return bool(VALID_DIRECT_TRK_CONTAINER_RE.match(candidate))


def is_currency_amount_text(text: str) -> bool:
    m = normalize(str(text).strip())
    if not m:
        return False
    if re.fullmatch(r"[\d.,]+\s*(usd|eur|gbp)", m):
        return True
    if re.fullmatch(r"(usd|eur|gbp)\s*[\d.,]+", m):
        return True
    return bool(re.search(r"\b(usd|eur|gbp)\b", m) and re.search(r"\d{3,}", m))


def is_container_type(measurement: str) -> bool:
    text = normalize_quotes(str(measurement)).strip()
    if not text or normalize(text) in GENERIC_UNIT_VALUES:
        return False
    if is_generic_per_container(measurement):
        return False
    if is_currency_amount_text(text):
        return False
    if is_pricing_or_tiered_text(text):
        return False
    return bool(CONTAINER_TYPE_RE.search(text))


def collect_tab_container_types(measurements: list[str]) -> list[str]:
    types: list[str] = []
    seen: set[str] = set()

    for meas in measurements:
        for part in split_container_measurements(str(meas).strip()):
            if not part or not is_valid_direct_trk_container(part):
                continue
            key = normalize(part)
            if key not in seen:
                seen.add(key)
                types.append(part)

    return types


def is_tiered_container_price(value) -> bool:
    if pd.isna(value):
        return False
    text = normalize_quotes(str(value))
    return bool(re.search(r"per\s+20\s*['']", text, re.I)) and (
        "days" in text.lower() or "onwards" in text.lower()
    )


def parse_tiered_container_prices(price_text: str) -> list[dict]:
    """Parse tiered per-container prices such as 'Days 4 to 7 $92 per 20' / $114 per 40'."""
    tiers: list[dict] = []
    text = normalize_quotes(str(price_text))

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        range_match = re.search(
            r"Days?\s+(\d+)\s+to\s+(\d+)\s+.*?"
            r"([\d.]+)\s+per\s+20\s*[''].*?"
            r"([\d.]+)\s+per\s+40",
            line,
            re.IGNORECASE,
        )
        if range_match:
            tiers.append({
                "day_from": int(range_match.group(1)),
                "day_to": int(range_match.group(2)),
                "price_20": round_up(range_match.group(3)),
                "price_40": round_up(range_match.group(4)),
            })
            continue

        onwards_match = re.search(
            r"Day\s+(\d+)\s+onwards.*?"
            r"([\d.]+)\s+per\s+20\s*[''].*?"
            r"([\d.]+)\s+per\s+40",
            line,
            re.IGNORECASE,
        )
        if onwards_match:
            tiers.append({
                "day_from": int(onwards_match.group(1)),
                "day_to": None,
                "price_20": round_up(onwards_match.group(2)),
                "price_40": round_up(onwards_match.group(3)),
            })

    return tiers


def tiered_rate_by(charge: str) -> str:
    n = normalize(charge)
    if "demurrage" in n:
        return "Quantity/Demurrage Day"
    if "detention" in n:
        return "Quantity/Detention Day"
    return "Quantity/Day"


def format_tiered_price(tiers: list[dict], container: str) -> str:
    """Format tiered prices for one container size (currency is in Currency column)."""
    price_key = "price_20" if container == "20'" else "price_40"
    lines: list[str] = []

    if tiers:
        lines.append(f"<{tiers[0]['day_from']} - 0")

    for tier in tiers:
        price = tier[price_key]
        price_str = int(price) if float(price) == int(float(price)) else price

        if tier["day_to"] is not None:
            lines.append(f"<={tier['day_to']} {price_str}")
        else:
            lines.append(f">={tier['day_from']} - {price_str}")

    return "\n".join(lines)


def build_tiered_container_rows(
    charge: str,
    sheet_name: str,
    entry: dict,
    flow_multiplier: str,
) -> list[dict]:
    tiers = parse_tiered_container_prices(entry["price_raw"])
    if not tiers:
        return []

    apply_if = build_apply_if(entry["lane_ids"])
    currency = entry["currency"] or "USD"
    rate_by = tiered_rate_by(charge)
    rows: list[dict] = []

    for container_label in ("20'", "40'"):
        rows.append({
            "Source tab": sheet_name,
            "Rate cost name": charge,
            "Rate Card Name": f"{charge.strip()} ({container_label})",
            "Multiplier": flow_multiplier,
            "Rate by": rate_by,
            "Currency": currency,
            "Price": format_tiered_price(tiers, container_label),
            "Max": "",
            "Apply if": apply_if,
        })

    return rows


def extract_trk_destination_city(*texts: object) -> str:
    for text in texts:
        if pd.isna(text):
            continue
        match = re.search(
            r"direct\s+trk\s+from\s+bremerhaven\s+to\s+(.+)",
            str(text),
            re.IGNORECASE,
        )
        if not match:
            continue
        city = match.group(1).strip()
        if city and city.lower() not in {"door", "t"}:
            return city
    return ""


def is_direct_trk_charge(charge: str) -> bool:
    text = normalize(charge)
    return "direct trk" in text and "bremerhaven" in text


def is_waiting_time_destination_charge(charge: str) -> bool:
    return "waiting time in destination" in normalize(charge)


def parse_waiting_time_destination_details(
    measurement: str,
    price_raw: object,
) -> tuple[str, str, str, object, str, object]:
    detail_text = " ".join(
        str(part).strip()
        for part in (measurement, price_raw)
        if part is not None and not pd.isna(part) and str(part).strip()
    )

    country = ""
    country_match = re.search(
        r"waiting time in\s+([A-Z]{2})\s*,",
        detail_text,
        re.IGNORECASE,
    )
    if country_match:
        country = country_match.group(1).upper()
    else:
        country_match = re.search(
            r"waiting time in\s+([A-Z]{2})\b",
            detail_text,
            re.IGNORECASE,
        )
        if country_match:
            country = country_match.group(1).upper()

    free_match = re.search(r"(\d+)\s*hours?\s*free", detail_text, re.IGNORECASE)
    free_hours = free_match.group(1) if free_match else "0"

    rate_card_name = (
        f"Waiting time in destination ({country}) - {free_hours} hours free units"
        if country
        else ""
    )

    currency, price = parse_currency_amount(detail_text)
    if price == "":
        currency, price = parse_currency_amount(price_raw)
    max_currency, max_price = extract_max_amount(price_raw)
    if max_price == "":
        _, max_price = extract_max_amount(detail_text)
        if max_price != "" and not currency:
            max_currency, _ = extract_max_amount(detail_text)
    rate_by = (
        "Quantity/30 minutes"
        if re.search(r"per\s+30\s*min", detail_text, re.IGNORECASE)
        else "Quantity/Hour"
    )
    return rate_card_name, rate_by, currency, price, max_currency, max_price


def extract_container_type_from_texts(*texts: object) -> str:
    for text in texts:
        if pd.isna(text):
            continue
        if is_currency_amount_text(str(text)) or is_pricing_or_tiered_text(str(text)):
            continue
        for part in split_container_measurements(str(text).strip()):
            if is_valid_direct_trk_container(part):
                return part
    return ""


def row_texts_for_container_search(entry: dict) -> list[str]:
    texts: list[str] = []
    for text in [entry.get("measurement", ""), *entry.get("row_texts", [])]:
        if pd.isna(text) or not str(text).strip():
            continue
        text_str = str(text).strip()
        if extract_trk_destination_city(text_str):
            continue
        if is_currency_amount_text(text_str):
            continue
        if is_generic_per_container(text_str):
            continue
        if is_pricing_or_tiered_text(text_str):
            continue
        texts.append(text_str)
    return texts


def containers_for_direct_trk_entry(
    entry: dict,
    tab_container_types: list[str],
    city_from_row: bool,
) -> list[str]:
    """Resolve container label(s) for one Direct TRK source row."""
    explicit = extract_container_type_from_texts(*row_texts_for_container_search(entry))
    if explicit:
        return [explicit]

    measurement = entry.get("measurement", "")
    if is_valid_direct_trk_container(measurement):
        return split_container_measurements(normalize_quotes(measurement))

    has_per_container = is_generic_per_container(measurement) or any(
        is_generic_per_container(str(text))
        for text in entry.get("row_texts", [])
    )
    if city_from_row and has_per_container and tab_container_types:
        return list(tab_container_types)

    return [""]


def extract_waiting_time_detail(measurement: str, row_texts: list[str]) -> str:
    for text in [measurement, *row_texts]:
        if pd.isna(text):
            continue
        text_str = str(text).strip()
        if re.search(r"waiting time in\s+[A-Z]{2}\b", text_str, re.IGNORECASE):
            return text_str
    return measurement


def build_waiting_time_destination_row(
    charge: str,
    sheet_name: str,
    entry: dict,
    flow_multiplier: str,
) -> dict | None:
    detail = extract_waiting_time_detail(
        entry.get("measurement", ""),
        entry.get("row_texts", []),
    )
    rate_card_name, rate_by, currency, price, max_currency, max_price = (
        parse_waiting_time_destination_details(
            detail,
            entry.get("price_raw"),
        )
    )
    if not rate_card_name:
        return None

    return {
        "Source tab": sheet_name,
        "Rate cost name": charge,
        "Rate Card Name": rate_card_name,
        "Multiplier": flow_multiplier,
        "Rate by": rate_by,
        "Currency": currency or entry.get("currency") or "USD",
        "Price": price,
        "Max": max_price if max_price != "" else "",
        "Apply if": build_apply_if(entry["lane_ids"]),
    }


def build_direct_trk_rows(
    charge: str,
    entries: list[dict],
    sheet_name: str,
    tab_container_types: list[str],
    flow_multiplier: str,
) -> list[dict]:
    rows: list[dict] = []
    last_city = ""

    for entry in entries:
        if is_tiered_container_price(entry.get("price_raw")):
            continue
        if is_pricing_or_tiered_text(str(entry.get("measurement", ""))):
            continue
        if is_pricing_or_tiered_text(str(entry.get("price_raw", ""))):
            continue

        city_from_row = extract_trk_destination_city(
            *entry.get("row_texts", []),
            entry.get("measurement", ""),
        )
        if city_from_row:
            city = city_from_row
            last_city = city
        elif last_city:
            city = last_city
        else:
            continue

        containers = containers_for_direct_trk_entry(
            entry, tab_container_types, bool(city_from_row)
        )
        if not containers:
            containers = [""]

        for container in containers:
            if container and not is_valid_direct_trk_container(container):
                continue
            if container:
                rate_card_name = f"Delivery Fee (to {city} ({container}))"
                rate_by = container
            else:
                rate_card_name = f"Delivery Fee (to {city})"
                rate_by = "ACC/Delivery Fee"

            max_currency, max_price = extract_max_from_entry(entry, entry.get("currency"))
            rows.append({
                "Source tab": sheet_name,
                "Rate cost name": charge,
                "Rate Card Name": rate_card_name,
                "Multiplier": flow_multiplier,
                "Rate by": rate_by,
                "Currency": entry.get("currency") or max_currency or "USD",
                "Price": entry.get("price"),
                "Max": max_price if max_price != "" else "",
                "Apply if": build_apply_if(entry["lane_ids"]),
            })

    return rows


def build_ccam_rows(
    charge: str,
    entries: list[dict],
    sheet_name: str,
    flow_multiplier: str,
) -> list[dict]:
    rows: list[dict] = []
    for entry in entries:
        max_currency, max_price = extract_max_from_entry(entry, entry.get("currency") or "USD")
        rows.append({
            "Source tab": sheet_name,
            "Rate cost name": charge,
            "Rate Card Name": "Additional CCAM Fee",
            "Multiplier": flow_multiplier,
            "Rate by": "ACC/Additional CCAM Fee",
            "Currency": entry.get("currency") or max_currency or "USD",
            "Price": entry.get("price"),
            "Max": max_price if max_price != "" else "",
            "Apply if": build_apply_if(entry["lane_ids"]),
        })
    return rows


def map_rate_card_and_rate_by(
    charge: str,
    measurement: str,
    sheet_name: str,
) -> tuple[str, str]:
    """Return (rate_card_name, rate_by) for a mapped accessorial charge."""
    n = normalize(charge)
    meas = normalize_quotes(measurement) if pd.notna(measurement) else ""
    country = extract_tab_country(sheet_name)

    def card(template: str) -> str:
        return template.format(measurement=meas, country=country)

    if "laden container" in n and "lift" in n:
        return card("Loading/Unloading/Re-entry Fee ({measurement})"), meas

    if "early gate in" in n or "nominate empty container" in n:
        return card("Early Gate In Fee ({measurement})"), meas

    if "milk run" in n:
        return card("Milk-Run Fee({measurement})"), meas

    if "cds liquidity" in n and "green" in n:
        return card("CDS liquidity at port (Green lane) ({measurement})"), meas

    if "cds liquidity" in n and "red" in n:
        return card("CDS liquidity at port ( Red lane) ({measurement})"), meas

    if n == "chb" or n.startswith("chb "):
        return "CHB Fee", "ACC/CHB Fee"

    if "depot storage" in n:
        return card("Depot Storage Fee({measurement})"), meas

    if "additional ccam" in n:
        return "Additional CCAM Fee", "ACC/Additional CCAM Fee"

    if "booking" in n and "amendment" not in n:
        return "Booking", "ACC/Booking"

    if "bill amendment" in n:
        return "Bill Amendment Fee", "Condition/Change of billing"

    if "direct trk" in f"{n} {normalize(meas)}" and "bremerhaven" in f"{n} {normalize(meas)}":
        city = extract_trk_destination_city(meas, charge)
        if city and is_container_type(meas):
            return f"Delivery Fee (to {city} ({meas}))", meas
        if city:
            return f"Delivery Fee (to {city})", "ACC/Delivery Fee"
        return "Delivery Fee", "ACC/Delivery Fee"

    if "staffing pickup" in n:
        return "Forklift Fee", "ACC/Forklift Fee"

    if "waiting time in china" in n or ("waiting time" in n and "china" in n):
        return "Waiting Time", "Quantity/Hour"

    if "waiting time in destination" in n:
        rate_card_name, rate_by, _, _, _, _ = parse_waiting_time_destination_details(
            meas, ""
        )
        return rate_card_name, rate_by

    if "t1 in cz" in n or n.startswith("t1 in cz"):
        return card("T1 (in CZ({measurement}))"), meas

    # Unmapped charges: no Rate Card Name, measurement goes to Rate by
    return "", meas


def build_apply_if(lane_ids: list[str]) -> str:
    lanes: list[str] = []
    seen: set[str] = set()
    for lane_id in lane_ids:
        if lane_id and lane_id not in seen:
            seen.add(lane_id)
            lanes.append(lane_id)
    if not lanes:
        return ""
    return f"LANE ID - Applies if Lane Id equals {', '.join(lanes)}"


def parse_lanes_from_apply_if(apply_if: str) -> list[str]:
    prefix = "LANE ID - Applies if Lane Id equals "
    if not apply_if or not apply_if.startswith(prefix):
        return []
    tail = apply_if[len(prefix):].strip()
    return [part.strip() for part in tail.split(",") if part.strip()]


def consolidate_accessorial_rows(rows: list[dict]) -> list[dict]:
    """Merge rows that differ only by Apply if lane id."""
    groups: dict[tuple, dict] = {}

    for row in rows:
        key = (
            row["Source tab"],
            row["Rate cost name"],
            row["Rate Card Name"],
            row["Multiplier"],
            row["Rate by"],
            row["Currency"],
            row["Price"],
            row.get("Max", ""),
        )
        lanes = parse_lanes_from_apply_if(row["Apply if"])

        if key not in groups:
            grouped = row.copy()
            grouped["_lanes"] = lanes
            groups[key] = grouped
            continue

        existing = groups[key]["_lanes"]
        seen = set(existing)
        for lane in lanes:
            if lane not in seen:
                existing.append(lane)
                seen.add(lane)

    result: list[dict] = []
    for grouped in groups.values():
        lanes = grouped.pop("_lanes")
        grouped["Apply if"] = build_apply_if(lanes)
        result.append(grouped)

    return result


def expand_entry_measurements(
    entry: dict,
    tab_container_types: list[str],
) -> list[dict]:
    """Expand one raw entry across container types and split combined measurements."""
    meas = entry["measurement"]

    if is_generic_per_container(meas):
        if tab_container_types:
            return [{**entry, "measurement": container} for container in tab_container_types]
        return [entry]

    parts = split_container_measurements(meas)
    if len(parts) > 1:
        return [{**entry, "measurement": part} for part in parts]

    return [entry]


def expand_charge_rows(
    row_entries: list[dict],
    tab_container_types: list[str],
) -> list[dict]:
    expanded: list[dict] = []
    for entry in row_entries:
        expanded.extend(expand_entry_measurements(entry, tab_container_types))
    return expanded


def duplicate_with_maersk(rows: list[dict]) -> list[dict]:
    result: list[dict] = []
    for row in rows:
        result.append(row)
        if row["Rate Card Name"]:
            maersk = row.copy()
            maersk["Rate Card Name"] = f"{row['Rate Card Name']} (Maersk)"
            result.append(maersk)
    return result


def transform_accessorial_df(
    df: pd.DataFrame,
    flow: str,
    sheet_name: str,
) -> pd.DataFrame:
    cols = list(df.columns)
    charge_col = find_col_by_aliases(cols, CHARGE_HEADER_ALIASES)
    lane_col = find_col_by_aliases(cols, LANE_HEADER_ALIASES)
    measurement_col = find_col_by_aliases(cols, MEASUREMENT_HEADER_ALIASES)
    comment_col = find_col_by_aliases(cols, COMMENT_HEADER_ALIASES)
    amount_col, header_currency = find_amount_currency_columns(cols)

    if not charge_col:
        print("  Warning: charge column not found — skipping tab")
        return pd.DataFrame(columns=ACCESSORIAL_OUTPUT_COLS)

    work = df.copy()
    work[charge_col] = work[charge_col].ffill()

    all_measurements: list[str] = []
    for _, row in work.iterrows():
        all_measurements.extend(row_cell_texts(row, cols))
    tab_container_types = collect_tab_container_types(all_measurements)

    cfg = FLOW_ACCESSORIAL_CONFIG.get(flow, FLOW_ACCESSORIAL_CONFIG["3"])
    multiplier = cfg["multiplier"]

    charge_groups: dict[str, list[dict]] = {}
    for _, row in work.iterrows():
        if row_has_hamida_comment(row, cols):
            continue

        charge = row.get(charge_col)
        comment = row.get(comment_col) if comment_col else None
        if is_skippable_row(charge, comment):
            continue

        measurement = row.get(measurement_col, "") if measurement_col else ""
        if pd.isna(measurement):
            measurement = ""

        currency, price = "", ""
        price_raw = row.get(amount_col) if amount_col else None
        if amount_col:
            currency, price = parse_currency_amount(price_raw, header_currency)

        row_texts = row_cell_texts(row, cols)
        container_type = extract_container_type_from_texts(measurement, *row_texts)

        charge_groups.setdefault(str(charge).strip(), []).append({
            "measurement": str(measurement).strip(),
            "currency": currency,
            "price": price,
            "price_raw": price_raw,
            "lane_ids": split_lane_ids(row.get(lane_col) if lane_col else ""),
            "comment": comment,
            "row_texts": row_texts,
            "container_type": container_type,
        })

    rows: list[dict] = []

    for charge, entries in charge_groups.items():
        if is_hamida_case_by_case_text(charge):
            continue
        if any(
            is_hamida_case_by_case_text(entry.get("comment"))
            or any(is_hamida_case_by_case_text(t) for t in entry.get("row_texts", []))
            for entry in entries
        ):
            continue

        if is_direct_trk_charge(charge):
            rows.extend(
                build_direct_trk_rows(
                    charge, entries, sheet_name, tab_container_types, multiplier
                )
            )
            continue

        if "additional ccam" in normalize(charge):
            rows.extend(build_ccam_rows(charge, entries, sheet_name, multiplier))
            continue

        if is_waiting_time_destination_charge(charge):
            for entry in entries:
                built = build_waiting_time_destination_row(
                    charge, sheet_name, entry, multiplier
                )
                if built:
                    rows.append(built)
            continue

        tiered_handled = False
        for entry in entries:
            if is_tiered_container_price(entry.get("price_raw")):
                tiered_handled = True
                rows.extend(
                    build_tiered_container_rows(
                        charge, sheet_name, entry, multiplier
                    )
                )
        if tiered_handled:
            continue

        expanded_entries = expand_charge_rows(entries, tab_container_types)

        for entry in expanded_entries:
            meas = normalize_quotes(entry["measurement"])
            rate_card_name, rate_by = map_rate_card_and_rate_by(charge, meas, sheet_name)
            apply_if = build_apply_if(entry["lane_ids"])
            max_currency, max_price = extract_max_from_entry(entry, entry.get("currency"))
            currency = entry["currency"] or max_currency

            rows.append({
                "Source tab": sheet_name,
                "Rate cost name": charge,
                "Rate Card Name": rate_card_name,
                "Multiplier": multiplier,
                "Rate by": rate_by,
                "Currency": currency,
                "Price": entry["price"],
                "Max": max_price if max_price != "" else "",
                "Apply if": apply_if,
            })

    rows = consolidate_accessorial_rows(rows)
    rows = duplicate_with_maersk(rows)
    return pd.DataFrame(rows, columns=ACCESSORIAL_OUTPUT_COLS)


def prompt_accessorial_sheets(xlsx: pd.ExcelFile) -> list[str] | None:
    answer = input("\nAdd accessorial costs? (y/n): ").strip().lower()
    if answer not in ("y", "yes"):
        return None

    print("\nAvailable sheets for accessorial costs:")
    for i, name in enumerate(xlsx.sheet_names, start=1):
        print(f"  {i}. {name}")

    choice = input("\nEnter sheet numbers or names (comma-separated): ").strip()
    if not choice:
        return None

    selected: list[str] = []
    for part in choice.split(","):
        part = part.strip()
        if not part:
            continue
        if part.isdigit():
            idx = int(part) - 1
            if 0 <= idx < len(xlsx.sheet_names):
                name = xlsx.sheet_names[idx]
                if name not in selected:
                    selected.append(name)
            continue

        if part in xlsx.sheet_names:
            if part not in selected:
                selected.append(part)
            continue

        match = next((n for n in xlsx.sheet_names if n.lower() == part.lower()), None)
        if match and match not in selected:
            selected.append(match)

    return selected or None


def load_accessorial_data(xlsx: pd.ExcelFile, sheet_names: list[str], flow: str) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for sheet in sheet_names:
        df = read_accessorial_sheet(xlsx, sheet)
        if df.empty:
            print(f"  No accessorial rows from '{sheet}'")
            continue

        transformed = transform_accessorial_df(df, flow, sheet)
        if not transformed.empty:
            frames.append(transformed)
            print(f"  Loaded accessorial tab '{sheet}' — {len(transformed)} rows")
        else:
            print(f"  No accessorial rows from '{sheet}'")

    if not frames:
        return pd.DataFrame(columns=ACCESSORIAL_OUTPUT_COLS)

    return pd.concat(frames, ignore_index=True)


def format_accessorial_sheet(ws, row_count: int) -> None:
    for col_idx, header in enumerate(ACCESSORIAL_OUTPUT_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_WRAP
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = ACCESSORIAL_COLUMN_WIDTHS.get(header, 14)

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    for row_idx in range(2, row_count + 2):
        price_cell = ws.cell(row=row_idx, column=ACCESSORIAL_OUTPUT_COLS.index("Price") + 1)
        line_count = str(price_cell.value or "").count("\n") + 1
        ws.row_dimensions[row_idx].height = max(18, 16 * line_count)

        for col_idx in range(1, len(ACCESSORIAL_OUTPUT_COLS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = FONT_NORMAL
            header = ACCESSORIAL_OUTPUT_COLS[col_idx - 1]
            cell.alignment = (
                ALIGN_LEFT_WRAP
                if header in (
                    "Rate cost name", "Rate Card Name", "Apply if", "Source tab", "Price", "Max"
                )
                else ALIGN_WRAP
            )


def append_accessorial_sheet(output_path: Path, df: pd.DataFrame) -> None:
    wb = load_workbook(output_path)
    if "Accessorial Costs" in wb.sheetnames:
        del wb["Accessorial Costs"]

    ws = wb.create_sheet("Accessorial Costs")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    format_accessorial_sheet(ws, len(df))
    wb.save(output_path)


def maybe_add_accessorial_costs(
    xlsx: pd.ExcelFile,
    output_path: Path | None,
    flow: str,
) -> None:
    if output_path is None or not output_path.exists():
        return

    sheet_names = prompt_accessorial_sheets(xlsx)
    if not sheet_names:
        return

    print(f"  Processing {len(sheet_names)} accessorial tab(s)...")
    accessorial_df = load_accessorial_data(xlsx, sheet_names, flow)
    if accessorial_df.empty:
        print("  No accessorial costs to add")
        return

    append_accessorial_sheet(output_path, accessorial_df)
    print(f"  Added 'Accessorial Costs' tab ({len(accessorial_df)} rows) → {output_path}")
