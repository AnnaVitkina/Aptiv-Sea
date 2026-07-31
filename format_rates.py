"""Transforms processed DataFrame rates into structured multi-header Excel.

Run: python format_rates.py
"""

# --- path bootstrap (must be before any local imports) ---
import sys, os
try:
    _scripts_dir = os.path.dirname(os.path.abspath(__file__))
except NameError:
    # exec(open(...)) in Colab — __file__ is not defined; find scripts dir
    _scripts_dir = None
    for _candidate in [
        "/content/Aptiv-Sea",
        "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team "
        "/Documents/AI Adoption RMT/RMT/Aptiv Sea",
    ]:
        if os.path.isfile(os.path.join(_candidate, "config.py")):
            _scripts_dir = _candidate
            break
    if _scripts_dir is None:
        _scripts_dir = os.getcwd()
if _scripts_dir not in sys.path:
    sys.path.insert(0, _scripts_dir)
# --------------------------------------------------------

import math
import pandas as pd
import re
from pathlib import Path
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from xlsx_to_df import main as load_excel
from process_df import process
from excel_layout import format_rates_workbook
from config import OUTPUT_DIR
from accessorial_costs import maybe_add_accessorial_costs


SHIPMENT_INFO_LABELS = [
    "Lane Id", "Origin City", "Origin postal code", "Origin country code",
    "Origin Country", "Origin region", "Destination city",
    "Destination postal code", "destination country code",
    "Destination country", "Destination region", "Payer region",
    "Origin port", "Destination port", "Service", "Carrier Name",
    "Valid from", "Valid to", "Equipment type",
]

SHIPMENT_INFO_FLOW2 = [
    "Lane Id", "Origin City", "Origin postal code", "Origin country code",
    "Destination city", "Destination postal code", "destination country code",
    "Equipment type",
]

SHIPMENT_INFO_FLOW3 = [
    "Lane Id", "Origin City", "Origin postal code", "Origin country code",
    "Destination city", "Destination postal code", "destination country code",
    "Carrier Name",
]

EQUIP_RENAME = {
    "20'": "20FT",
    "40'": "40FT",
}

EQUIP_RENAME_FLOW2 = {
    "20'": "FTL/Container 20FT",
    "40'": "FTL/Container 40FT",
    "40HC": "FTL/Container 40HC",
}

HEADER_ROWS = 6
BLOCK_ORDER = [
    "Origin Charges",
    "Main freight Charges",
    "Destination charges",
    "Other charges",
]


def normalize(s: str) -> str:
    return re.sub(r"[-_]", " ", str(s).strip().lower())


def find_col(columns: list[str], target: str) -> str | None:
    t = normalize(target)
    for c in columns:
        if normalize(c) == t:
            return c
    return None


def find_col_contains(columns: list[str], *keywords: str) -> str | None:
    """Find first column whose normalized name contains ALL given keywords."""
    for c in columns:
        cn = normalize(c)
        if all(kw in cn for kw in keywords):
            return c
    return None


def round_up(value, decimals=3):
    """Round a numeric value UP to the given decimal places."""
    if pd.isna(value):
        return value
    try:
        multiplier = 10 ** decimals
        return math.ceil(float(value) * multiplier) / multiplier
    except (ValueError, TypeError):
        return value


def is_false_like(value) -> bool:
    if pd.isna(value):
        return False
    if isinstance(value, bool):
        return value is False
    if isinstance(value, (int, float)):
        return float(value) == 0.0
    s = str(value).strip().lower()
    return s in {"false", "no", "n", "0"}


def is_true_like(value) -> bool:
    if pd.isna(value):
        return False
    if isinstance(value, bool):
        return value is True
    if isinstance(value, (int, float)):
        return float(value) != 0.0
    s = str(value).strip().lower()
    return s in {"true", "yes", "y", "1"}


def format_revision_date(value) -> str | None:
    if pd.isna(value):
        return None
    dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return None
    return dt.strftime("%d.%m.%Y")


def read_revision_validity(xlsx: pd.ExcelFile) -> tuple[str | None, str | None]:
    revision_tab = next(
        (s for s in xlsx.sheet_names if s.strip().lower() in {"revision", "revisions"}),
        None,
    )
    if not revision_tab:
        return None, None

    raw = pd.read_excel(xlsx, sheet_name=revision_tab, header=None)
    valid_from = None
    valid_to = None

    for r in range(len(raw)):
        for c in range(len(raw.columns)):
            cell = raw.iloc[r, c]
            text = str(cell).strip().lower()
            if text in {"", "nan"}:
                continue

            def pick_value() -> object:
                for nc in range(c + 1, len(raw.columns)):
                    v = raw.iloc[r, nc]
                    if not pd.isna(v) and str(v).strip() != "":
                        return v
                if ":" in str(cell):
                    return str(cell).split(":", 1)[1].strip()
                return None

            if valid_from is None and "valid from" in text:
                valid_from = format_revision_date(pick_value())
            if valid_to is None and ("until" in text or "valid to" in text):
                valid_to = format_revision_date(pick_value())

            if valid_from and valid_to:
                return valid_from, valid_to

    return valid_from, valid_to


def find_workbook_sheet(xlsx: pd.ExcelFile, *name_parts: str) -> str | None:
    """Find a sheet by exact name or by containing all name parts."""
    targets = {part.strip().lower() for part in name_parts}
    for sheet in xlsx.sheet_names:
        sheet_norm = sheet.strip().lower()
        if sheet_norm in targets:
            return sheet
    for sheet in xlsx.sheet_names:
        sheet_norm = sheet.strip().lower()
        if all(part in sheet_norm for part in targets):
            return sheet
    return None


def read_awarded_lane_ids(xlsx: pd.ExcelFile) -> set[str] | None:
    """Read Lane IDs from 'Awarded lanes to be implemented' when present."""
    sheet = find_workbook_sheet(xlsx, "awarded lanes to be implemented")
    if not sheet:
        return None

    df = pd.read_excel(xlsx, sheet_name=sheet)
    lane_col = (
        find_col(list(df.columns), "Lane Id")
        or find_col(list(df.columns), "Lane ID")
    )
    if not lane_col:
        print("  Warning: 'Awarded lanes to be implemented' tab found but Lane ID column is missing")
        return set()

    return {
        str(v).strip()
        for v in df[lane_col].dropna()
        if str(v).strip() and str(v).strip().lower() != "nan"
    }


def _row_lookup_key(row: pd.Series, key_columns: list[str]) -> tuple[str, ...]:
    return tuple(
        str(row[col]).strip() if col in row.index and pd.notna(row[col]) else ""
        for col in key_columns
    )


def read_source_column_values(
    file_path: Path,
    sheet_name: str,
    column_name: str,
    key_columns: list[str],
) -> dict[tuple[str, ...], object]:
    """Read exact workbook values for one column, keyed by shipment columns."""
    from openpyxl import load_workbook

    wb_formula = load_workbook(file_path, data_only=False, read_only=True)
    wb_values = load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]

        header_row = next(ws_formula.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map: dict[str, int] = {}
        for idx, name in enumerate(header_row):
            if name is None:
                continue
            col_map[normalize(str(name))] = idx

        value_idx = col_map.get(normalize(column_name))
        if value_idx is None:
            return {}

        key_idxs: list[int] = []
        for key_col in key_columns:
            idx = col_map.get(normalize(key_col))
            if idx is None:
                return {}
            key_idxs.append(idx)

        result: dict[tuple[str, ...], object] = {}
        rows_formula = ws_formula.iter_rows(min_row=2, values_only=False)
        rows_values = ws_values.iter_rows(min_row=2, values_only=True)
        for row_formula, row_values in zip(rows_formula, rows_values):
            key = tuple(
                str(row_formula[i].value).strip()
                if row_formula[i].value is not None else ""
                for i in key_idxs
            )
            if not any(key):
                continue

            cell = row_formula[value_idx]
            if cell.data_type == "f":
                val = row_values[value_idx]
            else:
                val = cell.value
            result[key] = val
        return result
    finally:
        wb_formula.close()
        wb_values.close()


def apply_exact_origin_handling_fee(
    df: pd.DataFrame,
    file_path: Path,
    sheet_name: str,
    key_labels: list[str] | None = None,
) -> pd.DataFrame:
    """Overwrite Origin handling fee min values with exact source workbook values."""
    column_name = "Pre-carriage Handling Charge min"
    dst_col = find_col(list(df.columns), column_name)
    if not dst_col:
        return df

    key_columns: list[str] = []
    for label in (key_labels or SHIPMENT_INFO_FLOW2):
        actual = find_col(list(df.columns), label)
        if actual:
            key_columns.append(actual)
    if not key_columns:
        return df

    source_values = read_source_column_values(
        file_path, sheet_name, column_name, key_columns
    )
    if not source_values:
        return df

    df = df.copy()
    new_values = []
    for _, row in df.iterrows():
        key = _row_lookup_key(row, key_columns)
        new_values.append(source_values.get(key, row[dst_col]))
    df[dst_col] = new_values
    return df


# ---------------------------------------------------------------------------
# Cost definition
# ---------------------------------------------------------------------------

@dataclass
class CostDef:
    display_name: str
    block: str
    min_col: str | None = None
    unit_col: str | None = None
    flat_col: str | None = None
    applies_if: str = "Applies if invoiced by Carrier"
    rate_by_override: str = ""
    row_filter_col: str | None = None
    row_filter_value: object = None
    _min_actual: str | None = field(default=None, init=False, repr=False)
    _unit_actual: str | None = field(default=None, init=False, repr=False)
    _flat_actual: str | None = field(default=None, init=False, repr=False)

    @property
    def has_min(self) -> bool:
        return self._min_actual is not None

    @property
    def has_punit(self) -> bool:
        return self._unit_actual is not None

    @property
    def has_flat(self) -> bool:
        return self.has_min or (self._flat_actual is not None and not self.has_punit)

    @property
    def num_cols(self) -> int:
        return 1 + int(self.has_flat) + int(self.has_punit)

    def get_rate_by(self) -> str:
        if self.rate_by_override:
            return self.rate_by_override
        ref = normalize(
            self._unit_actual or self._flat_actual
            or self.unit_col or self.flat_col or ""
        )
        if "per kg" in ref:
            base = "Weight/chargeable kg"
            dn = normalize(self.display_name)
            if "linehaul" in dn or "fuel" in dn:
                return f"{base}\n(Road)"
            return base
        if "cbm" in ref:
            return "Volume/cbm"
        return "Per shipment"

    def resolve(self, columns: list[str]) -> bool:
        if self.min_col:
            self._min_actual = find_col(columns, self.min_col)
        if self.unit_col:
            self._unit_actual = find_col(columns, self.unit_col)
        if self.flat_col:
            self._flat_actual = find_col(columns, self.flat_col)
        return any([self._min_actual, self._unit_actual, self._flat_actual])

    def is_all_zero(self, df: pd.DataFrame) -> bool:
        for col in [self._min_actual, self._unit_actual, self._flat_actual]:
            if col and col in df.columns:
                try:
                    vals = pd.to_numeric(df[col], errors="coerce").fillna(0)
                    if not vals.eq(0).all():
                        return False
                except Exception:
                    return False
        return True

    def used_columns(self) -> set[str]:
        return {c for c in [self._min_actual, self._unit_actual, self._flat_actual] if c}


# ---------------------------------------------------------------------------
# Flow 1 predefined costs
# ---------------------------------------------------------------------------

def get_flow1_costs() -> list[CostDef]:
    return [
        # --- Origin Charges ---
        CostDef("Pre-carriage Linehaul", "Origin Charges",
                min_col="Pre-carriage Linehaul Charge Min",
                unit_col="Pre-carriage Linehaul Charge per KG"),
        CostDef("Pre-carriage Fuel surcharge", "Origin Charges",
                min_col="Pre-carriage Fuel surcharge min",
                unit_col="Pre-carriage Fuel surcharge per KG"),
        CostDef("Origin handling fee", "Origin Charges",
                min_col="Pre-carriage Handling Charge min",
                unit_col="Pre-carriage Handling Charge per KG"),
        CostDef("Export customs clearance", "Origin Charges",
                flat_col="Origin Customs Clearance Charge"),

        # --- Main freight Charges ---
        CostDef("Transport cost", "Main freight Charges",
                min_col="LCL Charge Min", unit_col="LCL Charge/CBM",
                applies_if="Applies if:\n1. Equipment Type equals 'LTL/Standard'"),
        CostDef("BAF", "Main freight Charges",
                min_col="LCL BAF Min", unit_col="LCL BAF/CBM"),
        CostDef("Peak Season Surcharge", "Main freight Charges",
                unit_col="LCL PSS/CBM"),
        CostDef("Solas", "Main freight Charges",
                flat_col="VGA/Solas"),
        CostDef("Non-stackable shipment", "Main freight Charges",
                min_col="Non Stackable Fee Min", unit_col="Non Stackable Fee/CBM"),

        # --- Destination charges ---
        CostDef("On-carriage Linehaul", "Destination charges",
                min_col="On-carriage Linehaul Min",
                unit_col="On-carriage Linehaul per KG"),
        CostDef("On-carriage Fuel Surcharge", "Destination charges",
                min_col="On-carriage FSC Min",
                unit_col="On-carriage FSC per KG"),
        CostDef("Destination handling fee", "Destination charges",
                min_col="On-carriage Handling Charge Min",
                unit_col="On-carriage Handling Charge per KG"),
        CostDef("Import customs clearance", "Destination charges",
                flat_col="Destination Customs Clearance Fee per Shipment"),
    ]


# ---------------------------------------------------------------------------
# Storage (dynamic: one instance per unique "free days" value)
# ---------------------------------------------------------------------------

def build_storage_costs(df: pd.DataFrame, columns: list[str]) -> list[CostDef]:
    storage_col = find_col(columns, "Destination Storage per CBM per day (calendar)")
    if not storage_col:
        storage_col = find_col_contains(columns, "storage", "cbm", "day")
    free_col = find_col(columns, "Destination Storage free days")
    if not free_col:
        free_col = find_col_contains(columns, "storage", "free")
    if not storage_col:
        return []

    if free_col and free_col in df.columns:
        free_values = sorted(df[free_col].dropna().unique())
    else:
        free_values = [0]

    costs = []
    for fv in free_values:
        fd = int(fv) if pd.notna(fv) else 0
        costs.append(CostDef(
            display_name=f"Storage(Destination Storage per CBM per day ({fd} free))",
            block="Other charges",
            unit_col=storage_col,
            applies_if=(
                f"Applies if:\n1. Equipment Type equals 'LTL/Standard'\n"
                f"Applies if invoiced by Carrier\n"
                f"Comment: Free days: {fd}\n"
                f"Multiplier:\n• Quantity/Day (free units: {fd})"
            ),
            row_filter_col=free_col,
            row_filter_value=fd,
        ))
    return costs


# ---------------------------------------------------------------------------
# Unmatched cost columns → auto-generated CostDefs
# ---------------------------------------------------------------------------

def determine_block(name: str) -> str:
    n = normalize(name)
    if any(x in n for x in ("pre carriage", "origin")):
        return "Origin Charges"
    if any(x in n for x in ("on carriage", "destination")):
        return "Destination charges"
    if any(x in n for x in ("lcl", "main", "baf", "pss", "freight")):
        return "Main freight Charges"
    return "Other charges"


def find_unmatched_costs(df: pd.DataFrame, used: set[str],
                         shipment_cols: list[str]) -> list[CostDef]:
    ship_norm = {normalize(c) for c in shipment_cols}

    def should_skip(col_name: str) -> bool:
        cn = normalize(col_name)
        if cn == "currency":
            return True
        if "storage" in cn and "free" in cn:
            return True
        if "transit time" in cn:
            return True
        if "total price of container" in cn:
            return True
        return False

    unmatched = [
        c for c in df.columns
        if c not in used
        and normalize(c) not in ship_norm
        and not should_skip(c)
    ]
    if not unmatched:
        return []

    parsed_groups: dict[str, dict] = {}
    standalone: list[str] = []

    for col in unmatched:
        cn = normalize(col)
        if cn.endswith(" min"):
            base = cn[:-4].strip()
            parsed_groups.setdefault(base, {"min": None, "unit": None})
            parsed_groups[base]["min"] = col
        elif "per kg" in cn:
            base = re.sub(r"\s*per\s*kg\s*$", "", cn).strip()
            parsed_groups.setdefault(base, {"min": None, "unit": None})
            parsed_groups[base]["unit"] = col
        elif "/cbm" in cn or "per cbm" in cn:
            base = re.sub(r"\s*[/]?\s*(per\s*)?cbm.*$", "", cn).strip()
            parsed_groups.setdefault(base, {"min": None, "unit": None})
            parsed_groups[base]["unit"] = col
        else:
            standalone.append(col)

    used_in_groups: set[str] = set()
    results: list[CostDef] = []

    for _base, parts in parsed_groups.items():
        display = parts.get("min") or parts.get("unit") or _base
        display_clean = re.sub(
            r"\s*(min|per\s*kg|/cbm)$", "", display, flags=re.IGNORECASE
        ).strip()
        block = determine_block(display_clean)
        results.append(CostDef(
            display_clean, block,
            min_col=parts.get("min"), unit_col=parts.get("unit"),
        ))
        used_in_groups.update(c for c in parts.values() if c)

    for col in standalone:
        if col not in used_in_groups:
            results.append(CostDef(col, determine_block(col), flat_col=col))

    return results


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_shipment_cols(df: pd.DataFrame, labels: list[str] | None = None) -> list[str]:
    result = []
    for label in (labels or SHIPMENT_INFO_LABELS):
        actual = find_col(list(df.columns), label)
        if actual:
            result.append(actual)
    return result


LCL_CARRIER_COUNTRIES = frozenset({"DE", "HU", "IT", "PL", "PT", "RS"})


def assign_flow1_carrier_name(df: pd.DataFrame) -> pd.DataFrame:
    """Set Carrier Name for LCL (DHL) flow from destination country code."""
    dest_col = find_col(list(df.columns), "destination country code")
    if not dest_col:
        return df

    df = df.copy()
    codes = df[dest_col].astype(str).str.strip().str.upper()
    carrier = pd.Series("", index=df.index, dtype=str)

    for idx, code in codes.items():
        if not code or code == "NAN":
            continue
        if code == "MA":
            carrier.at[idx] = "DHL Logistic MA"
        elif code == "MK":
            carrier.at[idx] = "DHL Global Forwarding RS"
        elif code in LCL_CARRIER_COUNTRIES:
            carrier.at[idx] = f"DHL Global Forwarding {code}"

    df["Carrier Name"] = carrier
    return df


def assign_flow3_carrier_name(df: pd.DataFrame) -> pd.DataFrame:
    """Set Carrier Name in Flow 3 based on destination country code."""
    dest_col = find_col(list(df.columns), "destination country code")
    if not dest_col:
        return df

    df = df.copy()
    codes = df[dest_col].astype(str).str.strip().str.upper()
    carrier = pd.Series("", index=df.index, dtype=str)
    carrier[codes == "MA"] = "Maersk Logistics MA"
    carrier[codes.isin(["DE", "CZ"])] = "Damco CZ, Damco DE"
    df["Carrier Name"] = carrier
    return df


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def build_excel(df: pd.DataFrame, shipment_cols: list[str],
                costs: list[CostDef], currency_col: str | None,
                show_blocks: bool = True,
                standard_names: set[str] | None = None,
                awarded_lane_ids: set[str] | None = None,
                highlight_cells: list[tuple[int, int]] | None = None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"

    num_ship = len(shipment_cols)
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # --- shipment info header (row 6) ---
    for i, col_name in enumerate(shipment_cols, start=1):
        cell = ws.cell(row=HEADER_ROWS, column=i, value=col_name)
        cell.font = bold

    # --- determine column ranges for each cost and block ---
    rate_start = num_ship + 1
    cur = rate_start
    cost_ranges: list[tuple[CostDef, int, int]] = []
    block_ranges: dict[str, list[int]] = {}

    for cost in costs:
        s = cur
        e = cur + cost.num_cols - 1
        cost_ranges.append((cost, s, e))
        if cost.block not in block_ranges:
            block_ranges[cost.block] = [s, e]
        else:
            block_ranges[cost.block][1] = e
        cur = e + 1

    # --- Row 1: block names (merged) ---
    if show_blocks:
        for block in BLOCK_ORDER:
            if block in block_ranges:
                s, e = block_ranges[block]
                if s < e:
                    ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
                cell = ws.cell(row=1, column=s, value=block)
                cell.font = bold
                cell.alignment = wrap

    # --- Rows 2-6 per cost ---
    for cost, s, e in cost_ranges:
        # Row 2: cost display name
        if s < e:
            ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
        ws.cell(row=2, column=s, value=cost.display_name).font = bold

        # Row 3: applies if
        if s < e:
            ws.merge_cells(start_row=3, start_column=s, end_row=3, end_column=e)
        c3 = ws.cell(row=3, column=s, value=cost.applies_if)
        c3.alignment = wrap

        # Row 4: rate by
        if s < e:
            ws.merge_cells(start_row=4, start_column=s, end_row=4, end_column=e)
        c4 = ws.cell(row=4, column=s, value=cost.get_rate_by())
        c4.alignment = wrap

        # Row 5: MIN label (only in the Flat column when cost has a min source)
        if cost.has_flat and cost.has_min:
            ws.cell(row=5, column=s + 1, value="MIN")

        # Row 6: sub-column labels
        c = s
        ws.cell(row=HEADER_ROWS, column=c, value="Currency").font = bold
        c += 1
        if cost.has_flat:
            ws.cell(row=HEADER_ROWS, column=c, value="Flat").font = bold
            c += 1
        if cost.has_punit:
            ws.cell(row=HEADER_ROWS, column=c, value="p/unit").font = bold

    # --- Data rows ---
    for row_idx in range(len(df)):
        data_row = HEADER_ROWS + 1 + row_idx

        for col_idx, col_name in enumerate(shipment_cols, start=1):
            ws.cell(row=data_row, column=col_idx, value=df.iloc[row_idx][col_name])

        for cost, s, _e in cost_ranges:
            # Row filter: only write data if this row matches (Storage / Equipment)
            if cost.row_filter_col and cost.row_filter_col in df.columns:
                row_val = df.iloc[row_idx][cost.row_filter_col]
                expected = cost.row_filter_value
                try:
                    match = float(row_val) == float(expected)
                except (ValueError, TypeError):
                    match = str(row_val).strip() == str(expected).strip()
                if not match:
                    continue

            c = s

            # Currency
            if currency_col and currency_col in df.columns:
                ws.cell(row=data_row, column=c, value=df.iloc[row_idx][currency_col])
            c += 1

            # Flat
            if cost.has_flat:
                src = cost._min_actual if cost.has_min else cost._flat_actual
                if src and src in df.columns:
                    val = df.iloc[row_idx][src]
                    if normalize(cost.display_name) != "origin handling fee":
                        val = round_up(val)
                    ws.cell(row=data_row, column=c, value=val)
                c += 1

            # p/unit
            if cost.has_punit and cost._unit_actual and cost._unit_actual in df.columns:
                val = df.iloc[row_idx][cost._unit_actual]
                if normalize(cost.display_name) != "origin handling fee":
                    val = round_up(val)
                ws.cell(row=data_row, column=c, value=val)

    # --- Build cost_spans for formatting ---
    cost_spans = []
    for cost, s, e in cost_ranges:
        sub_columns = []
        c = s
        sub_columns.append(("Currency", None))
        c += 1
        if cost.has_flat:
            min_label = "MIN" if cost.has_min else None
            sub_columns.append(("Flat", min_label))
            c += 1
        if cost.has_punit:
            sub_columns.append(("p/unit", None))
        cost_spans.append((cost.display_name, s, e, sub_columns))

    data_start_row = HEADER_ROWS + 1
    data_row_count = len(df)
    unawarded_row_indices: set[int] = set()
    if awarded_lane_ids is not None:
        lane_col = next(
            (c for c in shipment_cols if normalize(c) == "lane id"),
            None,
        )
        if lane_col:
            for row_idx in range(len(df)):
                lane = str(df.iloc[row_idx][lane_col]).strip()
                if lane and lane.lower() != "nan" and lane not in awarded_lane_ids:
                    unawarded_row_indices.add(row_idx)

    format_rates_workbook(
        ws, shipment_cols, cost_spans,
        data_start_row, data_row_count,
        standard_names,
        unawarded_row_indices,
        highlight_cells,
    )

    return wb


# ---------------------------------------------------------------------------
# Flow 1
# ---------------------------------------------------------------------------

def flow_lcl(df_processed: pd.DataFrame, df_original: pd.DataFrame, file_path: Path):
    # Add Currency from original before any row filtering
    if not find_col(list(df_processed.columns), "Currency"):
        currency_src = find_col(list(df_original.columns), "Currency")
        if currency_src and currency_src in df_original.columns:
            df_processed = df_processed.copy()
            df_processed["Currency"] = df_original[currency_src].values

    equip_col = find_col(list(df_processed.columns), "Equipment type")
    if equip_col:
        before = len(df_processed)
        df_processed = df_processed[
            df_processed[equip_col].astype(str).str.contains("LCL|LTL", case=False, na=False)
        ].reset_index(drop=True)
        removed = before - len(df_processed)
        if removed:
            print(f"  Removed {removed} non-LCL rows ({len(df_processed)} remaining)")
    else:
        pass

    df_processed = assign_flow1_carrier_name(df_processed)
    all_cols = list(df_processed.columns)
    shipment_cols = get_shipment_cols(df_processed)

    # --- predefined costs ---
    costs = get_flow1_costs()
    resolved: list[CostDef] = []
    for cost in costs:
        if cost.resolve(all_cols):
            if not cost.is_all_zero(df_processed):
                resolved.append(cost)
            else:
                pass
        else:
            pass

    # --- storage ---
    storage_costs = build_storage_costs(df_processed, all_cols)
    for sc in storage_costs:
        if sc.resolve(all_cols) and not sc.is_all_zero(df_processed):
            resolved.append(sc)

    # --- collect used columns ---
    used: set[str] = set()
    for c in resolved:
        used.update(c.used_columns())

    # --- unmatched rate columns ---
    for uc in find_unmatched_costs(df_processed, used, shipment_cols):
        if uc.resolve(all_cols) and not uc.is_all_zero(df_processed):
            resolved.append(uc)

    # --- sort by block order (keep relative order within block) ---
    by_block: dict[str, list[CostDef]] = {b: [] for b in BLOCK_ORDER}
    for c in resolved:
        by_block.get(c.block, by_block["Other charges"]).append(c)
    sorted_costs: list[CostDef] = []
    for block in BLOCK_ORDER:
        sorted_costs.extend(by_block[block])

    currency_col = find_col(list(df_processed.columns), "Currency")

    print(f"  Building Excel — {len(shipment_cols)} shipment cols, {len(sorted_costs)} costs")

    # --- build & save ---
    std_names = {c.display_name for c in get_flow1_costs()}
    wb = build_excel(df_processed, shipment_cols, sorted_costs, currency_col,
                     standard_names=std_names)

    OUTPUT_DIR.mkdir(exist_ok=True)
    out = OUTPUT_DIR / f"{file_path.stem}_rates.xlsx"
    wb.save(out)
    print(f"  Saved → {out}")
    return out


# ---------------------------------------------------------------------------
# Flow 2 — FCL (no multiplier)
# ---------------------------------------------------------------------------

def get_flow2_costs() -> list[CostDef]:
    """Same costs as LCL but: only MIN values (Flat), no per-unit columns."""
    base_costs = [
        ("Pre-carriage Linehaul", "Pre-carriage Linehaul Charge Min"),
        ("Pre-carriage Fuel surcharge", "Pre-carriage Fuel surcharge min"),
        ("Origin handling fee", "Pre-carriage Handling Charge min"),
        ("Export customs clearance", "Origin Customs Clearance Charge"),
        ("Transport cost", "FCL Charge"),
        ("BAF", "FCL BAF Charge"),
        ("Peak Season Surcharge", "LCL PSS/CBM"),
        ("Solas", "VGA/Solas"),
        ("Non-stackable shipment", "Non Stackable Fee Min"),
        ("On-carriage Linehaul", "On-carriage Linehaul Min"),
        ("On-carriage Fuel Surcharge", "On-carriage FSC Min"),
        ("Destination handling fee", "On-carriage Handling Charge Min"),
        ("Import customs clearance", "Destination Customs Clearance Fee per Shipment"),
        ("ETS", "FCL ETSCharge"),
    ]
    costs = []
    for name, col in base_costs:
        applies = "" if name == "Transport cost" else f"Apply if: MEASUREMENT contains 'ACC/{name}' in any item"
        costs.append(CostDef(
            display_name=name,
            block="Other charges",
            flat_col=col,
            applies_if=applies,
            rate_by_override="Quantity/Percentage",
        ))
    return costs


def flow_qty_pct(df_processed: pd.DataFrame, df_original: pd.DataFrame,
                 xlsx: pd.ExcelFile, file_path: Path, source_sheet: str):
    df_original = df_original.copy()
    # Add Currency from original before any row filtering
    if not find_col(list(df_processed.columns), "Currency"):
        currency_src = find_col(list(df_original.columns), "Currency")
        if currency_src and currency_src in df_original.columns:
            df_processed = df_processed.copy()
            df_processed["Currency"] = df_original[currency_src].values

    # Optionally keep only Lane ID groups where at least one row has Existing Lane? = TRUE
    existing_lane_col = find_col(list(df_original.columns), "Existing Lane?")
    lane_id_col = find_col(list(df_original.columns), "Lane Id")
    remove_non_existing = input(
        "\nRemove rows where no row in the same Lane ID has 'Existing Lane?' = TRUE? (y/n): "
    ).strip().lower() in ("y", "yes")
    if remove_non_existing and existing_lane_col and existing_lane_col in df_original.columns and lane_id_col:
        before = len(df_processed)
        existing_true = df_original[existing_lane_col].map(is_true_like).fillna(False)
        lane_vals = df_original[lane_id_col]
        lane_has_id = lane_vals.notna() & lane_vals.astype(str).str.strip().ne("")
        lane_keep = existing_true.groupby(lane_vals).transform("any")
        keep_mask = pd.Series(True, index=df_original.index)
        keep_mask[lane_has_id] = lane_keep[lane_has_id]
        keep_mask[~lane_has_id] = existing_true[~lane_has_id]
        keep_mask = keep_mask.reindex(df_processed.index, fill_value=True)
        df_processed = df_processed.loc[keep_mask].reset_index(drop=True)
        df_original = df_original.loc[keep_mask].reset_index(drop=True)
        removed = before - len(df_processed)
        if removed:
            print(f"  Removed {removed} rows where no row in the same Lane ID has 'Existing Lane?' = TRUE")

    remove_lcl = input("\nRemove LCL shipments? (y/n): ").strip().lower()
    if remove_lcl in ("y", "yes"):
        equip_col = find_col(list(df_processed.columns), "Equipment type")
        if equip_col:
            before = len(df_processed)
            lcl_mask = ~df_processed[equip_col].astype(str).str.contains(
                "LCL|LTL", case=False, na=False
            )
            df_processed = df_processed.loc[lcl_mask].reset_index(drop=True)
            df_original = df_original.loc[lcl_mask].reset_index(drop=True)
            print(f"  Removed {before - len(df_processed)} LCL rows ({len(df_processed)} remaining)")
        else:
            pass

    df_processed = apply_exact_origin_handling_fee(
        df_processed, file_path, source_sheet, SHIPMENT_INFO_FLOW2
    )

    # Rename equipment type values
    equip_col_f2 = find_col(list(df_processed.columns), "Equipment type")
    if equip_col_f2:
        df_processed[equip_col_f2] = df_processed[equip_col_f2].astype(str).str.strip().replace(EQUIP_RENAME_FLOW2)

    all_cols = list(df_processed.columns)
    add_validity = input(
        "\nAdd shipment columns 'Valid from' and 'Valid to' from Revision tab? (y/n): "
    ).strip().lower() in ("y", "yes")
    shipment_labels = list(SHIPMENT_INFO_FLOW2)
    if add_validity:
        valid_from, valid_to = read_revision_validity(xlsx)
        if valid_from and valid_to:
            df_processed["Valid from"] = valid_from
            df_processed["Valid to"] = valid_to
            shipment_labels.extend(["Valid from", "Valid to"])
            print(f"  Added validity period from Revision tab: {valid_from} - {valid_to}")
        else:
            print("  Warning: Could not read both 'Valid from' and 'until' from Revision tab; columns not added")
    shipment_cols = get_shipment_cols(df_processed, shipment_labels)

    awarded_lane_ids = read_awarded_lane_ids(xlsx)
    if awarded_lane_ids is not None:
        print(
            f"  Awarded lanes tab found — {len(awarded_lane_ids)} lane IDs; "
            "rows with other Lane IDs will be highlighted in red"
        )

    costs = get_flow2_costs()
    resolved: list[CostDef] = []
    for cost in costs:
        if cost.resolve(all_cols):
            if not cost.is_all_zero(df_processed):
                resolved.append(cost)
            else:
                pass
        else:
            pass

    # Storage
    storage_costs = build_storage_costs(df_processed, all_cols)
    for sc in storage_costs:
        sc.applies_if = f"Apply if: MEASUREMENT contains 'ACC/{sc.display_name}' in any item"
        sc.rate_by_override = "Quantity/Percentage"
        sc.unit_col = None
        sc.flat_col = sc.unit_col
        if sc.resolve(all_cols) and not sc.is_all_zero(df_processed):
            resolved.append(sc)

    # Unmatched
    used: set[str] = set()
    for c in resolved:
        used.update(c.used_columns())
    per_kg_handling = find_col(all_cols, "Pre-carriage Handling Charge per KG")
    if per_kg_handling:
        used.add(per_kg_handling)

    for uc in find_unmatched_costs(df_processed, used, shipment_cols):
        uc.applies_if = f"Apply if: MEASUREMENT contains 'ACC/{uc.display_name}' in any item"
        uc.rate_by_override = "Quantity/Percentage"
        uc.unit_col = None
        if uc.resolve(all_cols) and not uc.is_all_zero(df_processed):
            resolved.append(uc)

    currency_col = find_col(list(df_processed.columns), "Currency")

    print(f"  Building Excel — {len(shipment_cols)} shipment cols, {len(resolved)} costs")

    std_names = {c.display_name for c in get_flow2_costs()}
    wb = build_excel(df_processed, shipment_cols, resolved, currency_col,
                     show_blocks=False, standard_names=std_names,
                     awarded_lane_ids=awarded_lane_ids)

    OUTPUT_DIR.mkdir(exist_ok=True)
    out = OUTPUT_DIR / f"{file_path.stem}_fcl.xlsx"
    wb.save(out)
    print(f"  Saved → {out}")
    return out


# ---------------------------------------------------------------------------
# Flow 3 — Multiplier (per equipment type)
# ---------------------------------------------------------------------------

FLOW3_BASE_COSTS = [
    ("Pre-carriage Linehaul", "Pre-carriage Linehaul Charge Min"),
    ("Pre-carriage Fuel surcharge", "Pre-carriage Fuel surcharge min"),
    ("Origin handling fee", "Pre-carriage Handling Charge min"),
    ("Export customs clearance", "Origin Customs Clearance Charge"),
    ("Transport cost", "FCL Charge"),
    ("BAF", "FCL BAF Charge"),
    ("Peak Season Surcharge", "LCL PSS/CBM"),
    ("Solas", "VGA/Solas"),
    ("Non-stackable shipment", "Non Stackable Fee Min"),
    ("On-carriage Linehaul", "On-carriage Linehaul Min"),
    ("On-carriage Fuel Surcharge", "On-carriage FSC Min"),
    ("Destination handling fee", "On-carriage Handling Charge Min"),
    ("Import customs clearance", "Destination Customs Clearance Fee per Shipment"),
    ("ETS", "FCL ETSCharge"),
]

FLOW3_DISPLAY_NAMES = {
    "Origin handling fee": "Pre-carriage Handling Charge",
    "BAF": "FCL BAF Charge",
}

FLOW3_COUNTRY_CODE_HIGHLIGHT_LANES = frozenset({"EMEA0369", "EMEA0325"})


def merge_by_lane(df: pd.DataFrame, shipment_cols: list[str],
                   equip_col: str, equip_types: list[str],
                   cost_specs: list[tuple[str, str]],
                   currency_col: str | None) -> tuple[pd.DataFrame, list[str]]:
    """Merge rows with the same Lane ID into one, spreading costs across equipment-type columns.

    Returns (merged_df, list of new column names created).
    """
    lane_col = find_col(list(df.columns), "Lane Id")
    if not lane_col:
        print("  Warning: Lane ID column not found — cannot merge rows")
        return df, []

    new_col_names: list[str] = []
    for cost_name, src_col in cost_specs:
        actual_src = find_col(list(df.columns), src_col)
        if not actual_src:
            continue
        for et in equip_types:
            col_key = f"{cost_name}__{et}"
            new_col_names.append(col_key)
            df[col_key] = None

    df[equip_col] = df[equip_col].astype(str).str.strip()

    for idx in df.index:
        et = df.at[idx, equip_col]
        for cost_name, src_col in cost_specs:
            actual_src = find_col(list(df.columns), src_col)
            if not actual_src:
                continue
            col_key = f"{cost_name}__{et}"
            if col_key in df.columns:
                df.at[idx, col_key] = df.at[idx, actual_src]

    group_cols = [c for c in shipment_cols if c in df.columns]
    if currency_col and currency_col in df.columns:
        group_cols.append(currency_col)

    agg = {}
    for col_key in new_col_names:
        if col_key in df.columns:
            agg[col_key] = "first"

    merged = df.groupby(group_cols, sort=False, dropna=False).agg(agg).reset_index()
    return merged, new_col_names


def build_flow3_costs(equip_types: list[str],
                      cost_specs: list[tuple[str, str]]) -> list[CostDef]:
    """Create CostDefs pointing to the pre-built merged columns (no row_filter)."""
    costs: list[CostDef] = []
    for name, _src_col in cost_specs:
        display_name = FLOW3_DISPLAY_NAMES.get(name, name)
        for et in equip_types:
            col_key = f"{name}__{et}"

            if name == "Transport cost":
                base_applies = ""
            else:
                base_applies = "Applies if invoiced by Carrier"

            multiplier = f"Multiplier:\nType/Container_{et}"
            applies = f"{base_applies}\n{multiplier}" if base_applies else multiplier

            costs.append(CostDef(
                display_name=f"{display_name}({et})",
                block="Other charges",
                flat_col=col_key,
                applies_if=applies,
                rate_by_override="Quantity/Percentage",
            ))
    return costs


def flow_multiplier(df_processed: pd.DataFrame, df_original: pd.DataFrame, file_path: Path):
    # Add Currency from original before any processing
    if not find_col(list(df_processed.columns), "Currency"):
        currency_src = find_col(list(df_original.columns), "Currency")
        if currency_src and currency_src in df_original.columns:
            df_processed = df_processed.copy()
            df_processed["Currency"] = df_original[currency_src].values

    equip_col = find_col(list(df_processed.columns), "Equipment type")
    if not equip_col:
        print("  Error: Equipment Type column not found")
        return None

    # Rename equipment type values in the data
    df_processed[equip_col] = df_processed[equip_col].astype(str).str.strip().replace(EQUIP_RENAME)

    equip_types = sorted(
        df_processed[equip_col].dropna().unique(), key=str
    )

    df_processed = assign_flow3_carrier_name(df_processed)
    shipment_cols = get_shipment_cols(df_processed, SHIPMENT_INFO_FLOW3)
    currency_col = find_col(list(df_processed.columns), "Currency")

    # Collect all cost specs (predefined + unmatched)
    all_cost_specs = list(FLOW3_BASE_COSTS)

    used_src = {normalize(col) for _, col in FLOW3_BASE_COSTS}
    ship_norm = {normalize(c) for c in shipment_cols}
    skip_norm = {"currency", "equipment type"}

    for col in df_processed.columns:
        cn = normalize(col)
        if cn in used_src or cn in ship_norm or cn in skip_norm:
            continue
        if cn.endswith(" min") or "per kg" in cn or "/cbm" in cn or "per cbm" in cn:
            continue
        if "transit time" in cn or "total price of container" in cn:
            continue
        actual = find_col(list(df_processed.columns), col)
        if actual:
            vals = pd.to_numeric(df_processed[actual], errors="coerce").fillna(0)
            if not vals.eq(0).all():
                all_cost_specs.append((col, col))

    # Merge rows by Lane ID, spreading costs per equipment type
    df_merged, new_cols = merge_by_lane(
        df_processed.copy(), shipment_cols, equip_col, equip_types,
        all_cost_specs, currency_col,
    )
    print(f"  Merged {len(df_processed)} rows → {len(df_merged)} lanes")

    # Build CostDefs pointing to the merged columns
    costs = build_flow3_costs(equip_types, all_cost_specs)
    resolved: list[CostDef] = []
    merged_cols = list(df_merged.columns)
    for cost in costs:
        if cost.resolve(merged_cols):
            if not cost.is_all_zero(df_merged):
                resolved.append(cost)

    print(f"  Building Excel — {len(shipment_cols)} shipment cols, {len(resolved)} costs")

    std_names = {
        f"{FLOW3_DISPLAY_NAMES.get(name, name)}({et})"
        for name, _ in FLOW3_BASE_COSTS
        for et in equip_types
    }

    highlight_cells: list[tuple[int, int]] = []
    lane_col = find_col(shipment_cols, "Lane Id")
    origin_cc_col = find_col(shipment_cols, "Origin country code")
    dest_cc_col = find_col(shipment_cols, "destination country code")
    if lane_col:
        for row_idx in range(len(df_merged)):
            lane = str(df_merged.iloc[row_idx][lane_col]).strip()
            if lane in FLOW3_COUNTRY_CODE_HIGHLIGHT_LANES:
                if origin_cc_col:
                    highlight_cells.append((row_idx, shipment_cols.index(origin_cc_col) + 1))
                if dest_cc_col:
                    highlight_cells.append((row_idx, shipment_cols.index(dest_cc_col) + 1))

    wb = build_excel(df_merged, shipment_cols, resolved, currency_col,
                     show_blocks=False, standard_names=std_names,
                     highlight_cells=highlight_cells)

    OUTPUT_DIR.mkdir(exist_ok=True)
    out = OUTPUT_DIR / f"{file_path.stem}_multiplier.xlsx"
    wb.save(out)
    print(f"  Saved → {out}")
    return out


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

FLOW_DESCRIPTIONS = {
    "1": "LCL flow — DHL Global Forwarding format. Grouped costs: "
         "Origin / Main Freight / Destination charges.",
    "2": "FCL (no multiplier) — DSV format. No grouped costs, "
         "each cost is pre-calculated if the measurement ACC/cost name is applied.",
    "3": "Multiplier — Damco/Maersk format. No grouped costs, "
         "each cost depends on equipment type, multiplier is applicable.",
}


def detect_flow(file_path: Path) -> str | None:
    name = file_path.stem.upper()
    if "MAERSK" in name or "DAMCO" in name:
        return "3"
    if "LCL" in name or "DHL" in name:
        return "1"
    if "DSV" in name:
        return "2"
    return None


if __name__ == "__main__":
    df_original, xlsx, file_path, source_sheet = load_excel()
    df_processed = process(df_original, xlsx)
    print(f"  Ready — {len(df_processed)} rows x {len(df_processed.columns)} cols")

    suggested = detect_flow(file_path)

    print("\nAvailable flows:")
    for key, desc in FLOW_DESCRIPTIONS.items():
        marker = "  <<<" if key == suggested else ""
        print(f"  {key}. {desc}{marker}")

    if suggested:
        confirm = input(
            f"\nDetected flow: {suggested} ({FLOW_DESCRIPTIONS[suggested]})\n"
            "Press Enter to confirm, or enter a different number (1/2/3): "
        ).strip()
        flow_choice = confirm if confirm else suggested
    else:
        flow_choice = input("\nCould not detect flow automatically. Enter choice (1/2/3): ").strip()

    output_path = None
    if flow_choice == "1":
        output_path = flow_lcl(df_processed, df_original, file_path)
    elif flow_choice == "2":
        output_path = flow_qty_pct(df_processed, df_original, xlsx, file_path, source_sheet)
    elif flow_choice == "3":
        output_path = flow_multiplier(df_processed, df_original, file_path)
    else:
        print(f"Invalid choice: {flow_choice}")
        sys.exit(1)

    maybe_add_accessorial_costs(xlsx, output_path, flow_choice)
