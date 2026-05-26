"""Apply Excel formatting to the rates layout workbook."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_ROWS = 6

SHIPMENT_BOLD_HEADERS = {
    "Lane Id", "Lane ID",
    "Origin Country",
    "Destination country", "Destination Country",
    "Service",
    "Carrier Name",
    "Valid from",
    "Valid to",
}

SHIPMENT_COLUMN_WIDTHS = {
    "Lane Id": 12, "Lane ID": 12,
    "Origin City": 16,
    "Origin postal code": 12, "Origin Postal Code": 12,
    "Origin country code": 10, "Origin Country code": 10, "Origin Country Code": 10,
    "Origin Country": 12,
    "Origin region": 14, "Origin Region": 14,
    "Destination city": 16, "Destination City": 16,
    "Destination postal code": 14, "Destination Postal Code": 14,
    "Destination country code": 12, "Destination Country code": 12, "Destination Country Code": 12,
    "Destination country": 12, "Destination Country": 12,
    "Destination region": 14, "Destination Region": 14,
    "Payer region": 14, "Payer Region": 14,
    "Origin port": 14,
    "Destination port": 14,
    "Service": 14,
    "Carrier Name": 18,
    "Valid from": 12,
    "Valid to": 12,
    "Equipment type": 14, "Equipment Type": 14,
}

RATE_SUBCOLUMN_WIDTHS = {
    "Currency": 10,
    "Flat": 12,
    "p/unit": 12,
}

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BLOCK_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

FONT_HEADER = Font(bold=True, size=10)
FONT_NORMAL = Font(bold=False)
ALIGN_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
ALIGN_LEFT_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")

CostSpan = tuple[str, int, int, list[tuple[str, str | None]]]


def format_rates_workbook(
    ws,
    shipment_columns: list[str],
    cost_spans: list[CostSpan],
    data_start_row: int,
    data_row_count: int,
    standard_display_names: set[str] | None = None,
) -> None:
    if standard_display_names is None:
        standard_display_names = set()
    shipment_width = len(shipment_columns)
    _apply_header_styles(ws, shipment_columns, shipment_width, cost_spans)
    _apply_data_fonts(ws, shipment_width, data_start_row, data_row_count)
    _apply_rate_highlights(
        ws, cost_spans, data_start_row, data_row_count, standard_display_names
    )
    _apply_column_widths(ws, shipment_columns, cost_spans)
    _apply_row_heights(ws, data_start_row, data_row_count)


def _apply_rate_highlights(
    ws,
    cost_spans: list[CostSpan],
    data_start_row: int,
    data_row_count: int,
    standard_display_names: set[str],
) -> None:
    """Green fill on cost name row and data values for non-standard costs."""
    last_row = data_start_row + max(data_row_count - 1, 0)
    for cost_name, start_col, end_col, _sub_columns in cost_spans:
        if cost_name in standard_display_names:
            continue
        highlight_rows = [2, *range(data_start_row, last_row + 1)]
        for row_idx in highlight_rows:
            for col_idx in range(start_col, end_col + 1):
                ws.cell(row=row_idx, column=col_idx).fill = GREEN_FILL


def _apply_header_styles(
    ws,
    shipment_columns: list[str],
    shipment_width: int,
    cost_spans: list[CostSpan],
) -> None:
    for col_idx, header in enumerate(shipment_columns, start=1):
        cell = ws.cell(row=HEADER_ROWS, column=col_idx)
        cell.font = FONT_HEADER if header in SHIPMENT_BOLD_HEADERS else FONT_NORMAL
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_WRAP

    for row_idx in range(1, HEADER_ROWS):
        for col_idx in range(shipment_width + 1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx == 1:
                cell.fill = BLOCK_FILL
            cell.alignment = ALIGN_LEFT_WRAP if row_idx in (3, 4) else ALIGN_WRAP

    for _cost_name, start_col, end_col, _sub_columns in cost_spans:
        for row_idx in range(2, HEADER_ROWS + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if row_idx == HEADER_ROWS:
                    cell.font = FONT_HEADER
                    cell.fill = HEADER_FILL
                cell.alignment = ALIGN_LEFT_WRAP if row_idx in (3, 4) else ALIGN_WRAP


def _apply_data_fonts(
    ws,
    shipment_width: int,
    data_start_row: int,
    data_row_count: int,
) -> None:
    last_row = data_start_row + max(data_row_count - 1, 0)
    for row_idx in range(data_start_row, last_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).font = FONT_NORMAL


def _apply_column_widths(
    ws,
    shipment_columns: list[str],
    cost_spans: list[CostSpan],
) -> None:
    for col_idx, name in enumerate(shipment_columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = SHIPMENT_COLUMN_WIDTHS.get(name, 14)

    for cost_name, start_col, end_col, sub_columns in cost_spans:
        for offset, (header, min_label) in enumerate(sub_columns):
            col_idx = start_col + offset
            letter = get_column_letter(col_idx)
            if min_label and min_label.startswith("<="):
                width = 10
            else:
                width = RATE_SUBCOLUMN_WIDTHS.get(header, 12)
                if header == "Flat" and len(cost_name) > 20:
                    width = 14
            ws.column_dimensions[letter].width = width


def _apply_row_heights(ws, data_start_row: int, data_row_count: int) -> None:
    for row_idx in range(1, HEADER_ROWS):
        ws.row_dimensions[row_idx].height = 38
    ws.row_dimensions[HEADER_ROWS].height = 24
    last_row = data_start_row + max(data_row_count - 1, 0)
    for row_idx in range(data_start_row, last_row + 1):
        ws.row_dimensions[row_idx].height = 16
